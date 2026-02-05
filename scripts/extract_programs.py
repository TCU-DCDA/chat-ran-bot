#!/usr/bin/env python3
"""
Extract program data from AddRan Excel spreadsheets into individual JSON files.

Each .xlsx in files/other_programs/ is a visual poster layout with data at specific
cell positions. This script reads those positions, extracts structured data, and
writes one JSON file per program to functions/program-data/.

Usage:
    python3 scripts/extract_programs.py
"""

import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# --- Paths ---
ROOT = Path(__file__).parent.parent
INPUT_DIR = ROOT / "files" / "other_programs"
OUTPUT_DIR = ROOT / "functions" / "program-data"
PROGRAMS_CSV = ROOT / "functions" / "programs.csv"

# Programs that already have dedicated data files — skip extraction
# Include the typo variant from the Excel filename
SKIP_NAMES = {"digital culture and data analytics", "digitial culture and data analytics"}

# Known abbreviation mappings for TCU AddRan programs
ABBREVIATIONS = {
    "Anthropology": "ANTH",
    "Comparative Race and Ethnic Studies": "CRES",
    "Comparative Race & Ethnic Studies": "CRES",
    "Creative Writing": "CRWT",
    "Criminology & Criminal Justice": "CRJU",
    "Criminology and Criminal Justice": "CRJU",
    "Criminal Justice": "CRJU",
    "Digital Culture and Data Analytics": "DCDA",
    "Economics": "ECON",
    "English": "ENGL",
    "General Studies": "GENS",
    "Geography": "GEOG",
    "History": "HIST",
    "International Economics": "INEQ",
    "Modern Language Studies": "MDLS",
    "Philosophy": "PHIL",
    "Political Science": "POSC",
    "Religion": "RELI",
    "Sociology": "SOCI",
    "Spanish and Hispanic Studies": "SPAN",
    "Spanish & Hispanic Studies": "SPAN",
    "Women and Gender Studies": "WGST",
    "Women & Gender Studies": "WGST",
}


def slugify(name, degree=""):
    """Convert program name + degree to a filename slug.
    'Economics' + 'BA' -> 'economics-ba'
    'History' + 'BA' -> 'history'  (no suffix if only one variant)
    """
    slug = name.lower()
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    slug = re.sub(r"\s+", "-", slug.strip())
    if degree and degree not in ("BA",):
        # Add degree suffix for BS/BGS variants
        slug = f"{slug}-{degree.lower()}"
    return slug


def get_merged_cell_value(ws, row, col):
    """Get cell value, resolving merged cell ranges."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def find_section_by_header(ws, header_text, col_range, row_range):
    """Search for header text in a cell range (case-insensitive partial match).
    Returns the row number where found, or None.
    """
    for row in range(row_range[0], row_range[1] + 1):
        for col in range(col_range[0], col_range[1] + 1):
            val = get_merged_cell_value(ws, row, col)
            if val and header_text.lower() in str(val).lower():
                return row
    return None


def lookup_program_url(program_name, degree):
    """Look up the TCU URL for a program from programs.csv."""
    if not PROGRAMS_CSV.exists():
        return ""
    name_lower = program_name.lower().strip()
    with open(PROGRAMS_CSV) as f:
        for line in f.readlines()[1:]:
            parts = line.strip().split(",")
            if len(parts) < 3:
                continue
            csv_name = parts[2].strip('"').strip().lower()
            if csv_name == name_lower:
                return parts[0].strip('"').strip()
            # Also try with "criminology" variants
            if "criminal" in name_lower and "criminal" in csv_name:
                return parts[0].strip('"').strip()
    return ""


def extract_program_name(ws):
    """Extract program name from cell H2."""
    return str(get_merged_cell_value(ws, 2, 8) or "").strip()


def extract_degree_type(ws):
    """Extract degree type from cell H4. Returns 'BA', 'BS', or 'BGS'."""
    raw = str(get_merged_cell_value(ws, 4, 8) or "").strip()
    if "science" in raw.lower():
        return "BS"
    elif "general" in raw.lower():
        return "BGS"
    elif "arts" in raw.lower():
        return "BA"
    return raw


def extract_total_hours(ws):
    """Extract total major hours from cell M8. '30 hours' -> 30.
    Falls back to scanning the description text for 'X hours' patterns.
    """
    raw = str(get_merged_cell_value(ws, 8, 13) or "").strip()
    match = re.search(r"(\d+)", raw)
    if match:
        return int(match.group(1))

    # Fallback: check N8
    raw2 = str(get_merged_cell_value(ws, 8, 14) or "").strip()
    match2 = re.search(r"(\d+)", raw2)
    if match2:
        return int(match2.group(1))

    # Fallback: scan descriptions for "complete X hours" or "X hours specifically"
    for row in [8, 32]:
        desc = str(get_merged_cell_value(ws, row, 1) or "")
        # Look for patterns like "43 hours specifically" or "complete 30 hours"
        hours_match = re.search(r"(\d{2,3})\s+hours?\s+(?:specifically|in\s+)", desc)
        if hours_match:
            return int(hours_match.group(1))
        hours_match = re.search(r"complete\s+(\d{2,3})\s+hours?", desc)
        if hours_match:
            return int(hours_match.group(1))

    return 0


def extract_descriptions(ws):
    """Extract description paragraphs from A8 and A32."""
    descs = []
    for row in [8, 32]:
        val = get_merged_cell_value(ws, row, 1)
        if val and str(val).strip():
            descs.append(str(val).strip())
    return descs


def extract_courses(ws):
    """Extract course listings from Column L (12), rows 10-40.

    Uses section headers ('Required', 'Elective') to separate categories.
    Identifies courses by the pattern: 3-4 uppercase letters + space + 4-5 digits.
    Also captures lettered category requirements (A., B., C., D.) for programs
    that describe requirements as categories rather than specific courses.
    """
    required = {"hours": 0, "courses": []}
    electives = {"hours": 0, "description": "", "courses": []}
    categories = []  # For lettered requirement categories (A., B., etc.)
    current_section = "required"

    for row in range(10, 40):
        # Check columns L (12) and M (13)
        val = get_merged_cell_value(ws, row, 12)
        if not val:
            val = get_merged_cell_value(ws, row, 13)
        if not val:
            continue
        text = str(val).strip()
        if not text:
            continue
        text_lower = text.lower()

        # Stop at non-course content
        if "foreign language" in text_lower and "req" in text_lower:
            break
        if "for class selections" in text_lower:
            break
        # Skip URLs but don't stop — courses may follow
        if text_lower.startswith("https://") or text_lower.startswith("http://"):
            continue

        # Detect section transitions
        if "elective" in text_lower and not re.search(r"[A-Z]{3,4}\s+\d{4,5}", text):
            current_section = "elective"
            hours_match = re.search(r"(\d+)\s*(?:hours|hrs)", text_lower)
            if hours_match:
                electives["hours"] = int(hours_match.group(1))
            if text and not electives["description"]:
                electives["description"] = text
            continue

        if "required" in text_lower and not re.search(r"[A-Z]{3,4}\s+\d{4,5}", text):
            current_section = "required"
            hours_match = re.search(r"(\d+)\s*(?:hours|hrs)", text_lower)
            if hours_match:
                required["hours"] = int(hours_match.group(1))
            # Also capture descriptive headers like "Survey Level - Three (3) courses - 9 hrs."
            if re.match(r"^[A-Z]", text) and "hrs" in text_lower and not text_lower.startswith("required"):
                categories.append(text)
            continue

        # Check if this is a course entry (CODE + number pattern)
        if re.search(r"[A-Z]{3,4}\s+\d{4,5}", text):
            if current_section == "required":
                required["courses"].append(text)
            else:
                electives["courses"].append(text)
        # Check for lettered category (A. ..., B. ..., etc.) or labeled requirements
        elif re.match(r"^[A-G][.:]", text):
            categories.append(text)
        # Capture descriptive requirement lines (e.g. "Survey Level - Three (3) courses - 9 hrs.")
        elif ("hrs" in text_lower or "courses" in text_lower) and re.match(r"^[A-Z]", text) and current_section == "required":
            categories.append(text)
        # Check for hours info in non-course lines
        elif current_section == "required" and required["hours"] == 0:
            hours_match = re.search(r"(\d+)\s*(?:hours|hrs)", text_lower)
            if hours_match:
                required["hours"] = int(hours_match.group(1))

    # If no specific courses were found but categories exist, use categories
    if not required["courses"] and categories:
        required["courses"] = categories

    return {"requiredCourses": required, "electiveCourses": electives}


def extract_career_options(ws):
    """Extract career options from Column U area.

    Searches for 'WHAT CAN I DO' header, then reads lettered items below it.
    """
    careers = []
    header_row = find_section_by_header(ws, "WHAT CAN I DO", (21, 28), (25, 50))
    if not header_row:
        return careers

    for row in range(header_row + 1, header_row + 15):
        # Check columns U (21) through AB (28)
        for col in [21, 24, 27, 28]:
            val = get_merged_cell_value(ws, row, col)
            if not val:
                continue
            text = str(val).strip()
            if not text or "what can" in text.lower() or "top career" in text.lower():
                continue
            # Stop if we've hit the contact section
            if "contact information" in text.lower():
                return careers
            # Skip items that look like contact data (names, phones, emails)
            if re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", text):
                return careers
            if re.search(r"\d{3}[-.\s]\d{3}[-.\s]\d{4}", text):
                return careers
            if text.lower().startswith("dr.") or "chair" in text.lower() or "director" in text.lower():
                return careers
            # Remove letter prefix (A., B., etc.)
            cleaned = re.sub(r"^[A-H]\.\s*", "", text).strip()
            if cleaned and cleaned not in careers:
                careers.append(cleaned)

    return careers


def extract_contacts(ws):
    """Extract contact information from Column U area.

    Searches for 'CONTACT INFORMATION' header, then parses role/name/phone/email/office blocks.
    """
    contacts = []
    header_row = find_section_by_header(ws, "CONTACT INFORMATION", (21, 28), (35, 65))
    if not header_row:
        return contacts

    role_keywords = [
        "chair", "director", "consultant", "advisor", "coordinator",
    ]

    i = header_row + 1
    while i <= min(header_row + 25, 70):
        val = get_merged_cell_value(ws, i, 21)
        if not val:
            i += 1
            continue
        text = str(val).strip()
        text_lower = text.lower()

        # Check if this row is a role title
        is_role = any(kw in text_lower for kw in role_keywords)
        if is_role:
            contact = {"role": text}
            # Read next rows for name, phone, email, office
            for j in range(i + 1, min(i + 6, 71)):
                sub_val = get_merged_cell_value(ws, j, 21)
                if not sub_val:
                    continue
                sub_text = str(sub_val).strip()
                if not sub_text:
                    continue

                # Check if we've hit the next role
                if any(kw in sub_text.lower() for kw in role_keywords):
                    break

                # Detect email
                email_match = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", sub_text)
                if email_match:
                    contact["email"] = email_match.group(0)
                    continue

                # Detect phone (817-xxx-xxxx pattern)
                phone_match = re.search(r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", sub_text)
                if phone_match:
                    contact["phone"] = phone_match.group(0)
                    continue

                # First unidentified text = name, second = office
                if "name" not in contact:
                    contact["name"] = sub_text
                elif "office" not in contact:
                    contact["office"] = sub_text

            contacts.append(contact)
            i += 5
        else:
            i += 1

    return contacts


def extract_internship(ws):
    """Extract internship info from Column U, rows 7-25."""
    internship = {}
    # Look for internship header
    header_row = find_section_by_header(ws, "INTERNSHIP", (21, 28), (5, 28))
    if not header_row:
        return internship

    desc_parts = []
    for row in range(header_row + 1, min(header_row + 10, 30)):
        val = get_merged_cell_value(ws, row, 21)
        if not val:
            continue
        text = str(val).strip()
        if not text:
            continue
        # Stop if we hit a new major section
        if "WHAT CAN I DO" in text.upper() or "CONTACT" in text.upper():
            break
        # Skip sub-section headers that are just "Inclusive Excellence" etc.
        if "inclusive excellence" in text.lower():
            break
        desc_parts.append(text)

    if desc_parts:
        internship["description"] = " ".join(desc_parts)

    return internship


def extract_single_program(xlsx_path):
    """Extract all data from a single Excel file and return a dict."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Use the first sheet (main program sheet, not the grade lookup)
    ws = wb.worksheets[0]

    name = extract_program_name(ws)
    if not name:
        name = xlsx_path.stem.split("(")[0].strip()

    degree = extract_degree_type(ws)
    abbr = ABBREVIATIONS.get(name, "")

    program = {
        "name": name,
        "abbreviation": abbr,
        "degree": degree,
        "totalHours": extract_total_hours(ws),
        "url": lookup_program_url(name, degree),
        "descriptions": extract_descriptions(ws),
        "requirements": extract_courses(ws),
        "careerOptions": extract_career_options(ws),
        "contacts": extract_contacts(ws),
        "internship": extract_internship(ws),
    }

    wb.close()
    return program


def determine_slug(name, degree, all_files):
    """Determine output filename slug, adding degree suffix when needed for disambiguation."""
    # Check if multiple files share the same base program name
    base_name = name.lower().strip()
    variants = [f for f in all_files if f.stem.split("(")[0].strip().lower() == base_name
                or name.lower() in f.stem.lower()]

    # Count how many Excel files map to this program name
    if len(variants) > 1 or degree in ("BS", "BGS"):
        slug = slugify(name) + f"-{degree.lower()}"
    else:
        slug = slugify(name)
    return slug


def main():
    if not INPUT_DIR.exists():
        print(f"ERROR: Input directory not found: {INPUT_DIR}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted([f for f in INPUT_DIR.glob("*.xlsx") if not f.name.startswith("~$")])
    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in {INPUT_DIR}")
        sys.exit(1)

    print(f"Found {len(xlsx_files)} Excel files to process.\n")

    results = []
    errors = []

    for xlsx_path in xlsx_files:
        print(f"Processing: {xlsx_path.name}")
        try:
            program = extract_single_program(xlsx_path)

            # Skip programs with dedicated data files
            if program["name"].lower().strip() in SKIP_NAMES:
                print(f"  SKIPPED (already has dedicated data file)")
                continue

            slug = determine_slug(program["name"], program["degree"], xlsx_files)
            output_path = OUTPUT_DIR / f"{slug}.json"

            with open(output_path, "w") as f:
                json.dump(program, f, indent=2)

            print(f"  -> {output_path.name}")
            results.append({
                "file": xlsx_path.name,
                "output": output_path.name,
                "name": program["name"],
                "degree": program["degree"],
                "hours": program["totalHours"],
                "courses": len(program["requirements"].get("requiredCourses", {}).get("courses", [])),
                "careers": len(program["careerOptions"]),
                "contacts": len(program["contacts"]),
            })
        except Exception as e:
            print(f"  ERROR: {e}")
            errors.append({"file": xlsx_path.name, "error": str(e)})

    # Summary
    print(f"\n{'='*60}")
    print(f"EXTRACTION SUMMARY")
    print(f"{'='*60}")
    print(f"Processed: {len(results)}/{len(xlsx_files)}")
    print(f"Errors:    {len(errors)}")

    if errors:
        print(f"\nFailed files:")
        for e in errors:
            print(f"  - {e['file']}: {e['error']}")

    print(f"\n{'Name':<40s} {'Deg':>3s} {'Hrs':>3s} {'Crs':>3s} {'Car':>3s} {'Con':>3s}")
    print("-" * 60)
    for r in results:
        print(f"  {r['name']:<38s} {r['degree']:>3s} {r['hours']:>3d} {r['courses']:>3d} {r['careers']:>3d} {r['contacts']:>3d}")


if __name__ == "__main__":
    main()
